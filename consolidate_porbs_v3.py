#!/usr/bin/env python3
"""
PORB Excel File Consolidation Script v3
- Updated for new <Status>_PORB_<SPCode>_<Center>[_suffix].xlsx file format
- Accepts any prefix/status (Draft_, Pending_, Final_, etc.)
- Handles trailing suffixes after center name (e.g., _error_outline, _v2)
- Handles SP code to program name mapping
- Supports 4 new sheets: Outcomes, Synergy Programs, Countries of Implementation, Location of Benefit
- Center column is now embedded in data (not inserted by script)
- Drops internal ID columns (id, standerd_cross_cutting_id)
- Supports --input-dir command-line argument
- Creates per-program summary files and a master file
"""

import argparse
import os
import re
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Default directories
DEFAULT_INPUT_DIR = os.path.join(SCRIPT_DIR, "PORBs")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "SummaryPORBs_v3")

# SP code to program name mapping
SP_CODE_MAP = {
    "SP01": "Breeding for Tomorrow",
    "SP02": "Sustainable Farming",
    "SP03": "Sustainable Animal and Aquatic Foods",
    "SP04": "Multifunctional Landscapes",
    "SP05": "Better Diets and Nutrition",
    "SP06": "Climate Action",
    "SP07": "Policy Innovations",
    "SP08": "Food Frontiers and Security",
    "SP09": "Scaling for Impact",
    "SP10": "Gender Equality and Inclusion",
    "SP11": "Capacity Sharing",
    "SP12": "Digital Transformation",
    "SP13": "Genebank",
}

PROGRAM_NAMES = list(SP_CODE_MAP.values())

# Known CGIAR center names (used to extract center from filenames with suffixes)
KNOWN_CENTERS = [
    "AfricaRice",
    "Bioversity (Alliance)",
    "Bioversity_Alliance",
    "CIAT (Alliance)",
    "CIAT_Alliance",
    "CIMMYT",
    "CIP",
    "ICARDA",
    "ICRISAT",
    "IFPRI",
    "IITA",
    "ILRI",
    "IRRI",
    "IWMI",
    "SO",
    "WorldFish",
]

# All data sheets to consolidate (in desired output order)
DATA_SHEETS = [
    "HLO",
    "Partners",
    "W3-Bilateral",
    "MELIA",
    "Cross Cutting",
    "Outcomes",
    "Synergy Programs",
    "Countries of Implementation",
    "Location of Benefit",
]

# Columns that might have merged cells and need forward-fill
MERGE_FILL_COLS = {
    "HLO": ["AOW", "High Level Output"],
    "Partners": ["AOW"],
    "W3-Bilateral": ["Center", "Project title"],
    "MELIA": ["AOW"],
    "Cross Cutting": ["AOW"],
    "Outcomes": ["AOW", "Outcome"],
    "Synergy Programs": ["AOW"],
    "Countries of Implementation": ["AOW"],
    "Location of Benefit": ["AOW"],
}

# Internal ID columns to drop from output
DROP_COLUMNS = {"id", "standerd_cross_cutting_id", "aow_id"}

# HLO sheet has a two-row header - define proper column names
HLO_COLUMNS = [
    "AOW",
    "Center",
    "High Level Output",
    "Description",
    "KPI Type",
    "Country(ies) of implementation",
    "Target",
    "Budget (USD)",
    "Assumption",
    "id",
]


def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Consolidate Draft PORB Excel files into summary workbooks."
    )
    parser.add_argument(
        "--input-dir",
        default=DEFAULT_INPUT_DIR,
        help="Directory containing Draft_PORB_*.xlsx files (default: PORBs subfolder)",
    )
    return parser.parse_args()


def extract_program_center(filename):
    """Extract program name and center name from <Status>_PORB_<SPCode>_<Center>[_suffix].xlsx.

    Accepts any prefix/status before _PORB_ (Draft, Pending, Final, etc.)
    and handles trailing suffixes after the center name by matching against
    known CGIAR centers. If no known center matches, falls back to using
    everything after the SP code as the center name.

    Examples:
        Draft_PORB_SP01_AfricaRice.xlsx           → ("Breeding for Tomorrow", "AfricaRice")
        Pending_PORB_SP07_IFPRI.xlsx              → ("Policy Innovations", "IFPRI")
        Draft_PORB_SP04_AfricaRice_error_outline.xlsx → ("Multifunctional Landscapes", "AfricaRice")
        Final_PORB_SP12_CIMMYT_v2.xlsx            → ("Digital Transformation", "CIMMYT")
    """
    basename = filename.replace(".xlsx", "")
    match = re.match(r"^.+_PORB_(SP\d{2})_(.+)$", basename)
    if not match:
        return None, None
    sp_code = match.group(1)
    rest = match.group(2)  # e.g. "AfricaRice_error_outline" or "IFPRI"
    program = SP_CODE_MAP.get(sp_code)
    if not program:
        print(f"  Warning: Unknown SP code '{sp_code}' in file {filename}")
        return None, None

    # Try to match a known center name at the start of 'rest'
    # Sort by length descending so longer names match first (e.g. "CIAT (Alliance)" before "CIAT")
    for center_name in sorted(KNOWN_CENTERS, key=len, reverse=True):
        # Normalize for comparison (handle spaces, parens, underscores)
        normalized_rest = rest.replace("_", " ").lower()
        normalized_center = center_name.replace("_", " ").lower()
        if normalized_rest == normalized_center or normalized_rest.startswith(normalized_center + " "):
            return program, center_name
        # Also check with underscore separators (as filenames use underscores)
        center_underscored = center_name.replace(" ", "_")
        if rest == center_underscored or rest.startswith(center_underscored + "_"):
            return program, center_name

    # Fallback: use the full 'rest' as center name (handles unknown centers)
    # But try to strip common suffixes like _error_outline, _v2, etc.
    center = rest.split("_")[0] if "_" in rest else rest
    print(f"  Note: Center '{center}' (from {filename}) not in known centers list — using as-is")
    return program, center


def drop_id_columns(df):
    """Drop internal ID columns from a dataframe."""
    cols_to_drop = [c for c in df.columns if c.lower().strip() in DROP_COLUMNS or c in DROP_COLUMNS]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    return df


def clean_dataframe(df, sheet_name):
    """Clean dataframe - handle merged cells, remove empty rows and subtotals."""
    if df.empty:
        return df

    # Forward-fill merged cell columns
    if sheet_name in MERGE_FILL_COLS:
        for col in MERGE_FILL_COLS[sheet_name]:
            if col in df.columns:
                df[col] = df[col].ffill()

    # Remove rows that are completely empty
    df = df.dropna(how="all")

    # Remove subtotal rows
    for col in df.columns:
        if df[col].dtype == object:
            mask = df[col].astype(str).str.contains(
                r"subtotal|Subtotal|SUBTOTAL", na=False, regex=True
            )
            df = df[~mask]

    return df.reset_index(drop=True)


def read_hlo_sheet(filepath):
    """Read HLO sheet with special handling for two-row headers."""
    df = pd.read_excel(filepath, sheet_name="HLO", header=None)

    # Skip the first two header rows
    df = df.iloc[2:].reset_index(drop=True)

    # Assign proper column names
    if len(df.columns) == len(HLO_COLUMNS):
        df.columns = HLO_COLUMNS
    elif len(df.columns) > len(HLO_COLUMNS):
        df.columns = HLO_COLUMNS + [
            f"Extra_{i}" for i in range(len(df.columns) - len(HLO_COLUMNS))
        ]
    else:
        df.columns = HLO_COLUMNS[: len(df.columns)]

    return df


def read_and_consolidate_sheet(files, sheet_name, porb_dir):
    """Read a specific sheet from all files and consolidate."""
    all_data = []

    for filename in files:
        program, center = extract_program_center(filename)
        if not program:
            continue

        filepath = os.path.join(porb_dir, filename)
        try:
            # Special handling for HLO sheet with two-row headers
            if sheet_name == "HLO":
                df = read_hlo_sheet(filepath)
            else:
                df = pd.read_excel(filepath, sheet_name=sheet_name)

            df = clean_dataframe(df, sheet_name)
            df = drop_id_columns(df)

            if not df.empty:
                # Insert Program/Accelerator as first column (Center is already in data)
                df.insert(0, "Program/Accelerator", program)
                all_data.append(df)
        except ValueError:
            # Sheet does not exist in the file - skip silently
            pass
        except Exception as e:
            print(f"  Warning: Error reading '{sheet_name}' from {filename}: {e}")

    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        return combined
    return pd.DataFrame()


def read_anaplan_sheet(files, porb_dir):
    """Read and consolidate Anaplan sheets."""
    all_data = []

    for filename in files:
        program, center = extract_program_center(filename)
        if not program:
            continue

        filepath = os.path.join(porb_dir, filename)
        try:
            df = pd.read_excel(filepath, sheet_name="Anaplan")
            df = df.dropna(how="all")
            df = drop_id_columns(df)

            if "Main Accounts" in df.columns:
                mask = df["Main Accounts"].astype(str).str.contains(
                    r"Subtotal|subtotal", na=False, regex=True
                )
                df = df[~mask]

            if not df.empty:
                df.insert(0, "Center", center)
                df.insert(0, "Program/Accelerator", program)
                all_data.append(df)
        except ValueError:
            pass
        except Exception as e:
            print(f"  Warning: Error reading Anaplan from {filename}: {e}")

    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        return combined
    return pd.DataFrame()


def style_header(ws, num_cols):
    """Apply professional styling to header row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def auto_adjust_columns(ws, df):
    """Auto-adjust column widths based on content."""
    for idx, col in enumerate(df.columns, 1):
        max_length = len(str(col))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except Exception:
                    pass
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max_length + 2


def write_consolidated_file(output_path, sheet_data_dict, title_prefix=""):
    """Write consolidated data to Excel file with multiple sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    sheets_written = 0
    for sheet_name, df in sheet_data_dict.items():
        if df.empty:
            continue

        # Excel sheet names are limited to 31 characters
        clean_name = sheet_name[:31].replace("/", "-")
        ws = wb.create_sheet(title=clean_name)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1:
                    cell.font = Font(name="Arial", size=10)

        style_header(ws, len(df.columns))
        auto_adjust_columns(ws, df)
        ws.freeze_panes = "A2"
        sheets_written += 1

    if sheets_written > 0:
        wb.save(output_path)
        print(f"  Saved: {output_path} ({sheets_written} sheets)")
    else:
        print(f"  Skipped: {output_path} (no data)")


def create_program_summary(program, files, porb_dir, output_dir):
    """Create summary file for a single program."""
    program_files = [f for f in files if extract_program_center(f)[0] == program]

    if not program_files:
        print(f"  No files found for {program}")
        return

    centers = [extract_program_center(f)[1] for f in program_files]
    print(f"\nProcessing: {program} ({len(program_files)} centers: {', '.join(sorted(centers))})")

    sheet_data = {}

    for sheet_name in DATA_SHEETS:
        df = read_and_consolidate_sheet(program_files, sheet_name, porb_dir)
        if not df.empty:
            # Drop Program/Accelerator for per-program files (redundant)
            if "Program/Accelerator" in df.columns:
                df = df.drop(columns=["Program/Accelerator"])
            sheet_data[sheet_name] = df
            print(f"    {sheet_name}: {len(df)} rows")
        else:
            print(f"    {sheet_name}: (empty)")

    # Anaplan sheet
    df_anaplan = read_anaplan_sheet(program_files, porb_dir)
    if not df_anaplan.empty:
        if "Program/Accelerator" in df_anaplan.columns:
            df_anaplan = df_anaplan.drop(columns=["Program/Accelerator"])
        sheet_data["Anaplan"] = df_anaplan
        print(f"    Anaplan: {len(df_anaplan)} rows")
    else:
        print(f"    Anaplan: (empty)")

    safe_name = re.sub(r"[^\w\s-]", "", program).strip()
    output_path = os.path.join(output_dir, f"PORB_Consolidated_{safe_name}.xlsx")
    write_consolidated_file(output_path, sheet_data, program)


def create_master_file(files, porb_dir, output_dir):
    """Create master file with all programs combined."""
    print("\n" + "-" * 60)
    print("Creating MASTER file (all programs)...")
    print("-" * 60)

    sheet_data = {}

    for sheet_name in DATA_SHEETS:
        df = read_and_consolidate_sheet(files, sheet_name, porb_dir)
        if not df.empty:
            sheet_data[sheet_name] = df
            print(f"    {sheet_name}: {len(df)} rows")
        else:
            print(f"    {sheet_name}: (empty)")

    # Anaplan sheet
    df_anaplan = read_anaplan_sheet(files, porb_dir)
    if not df_anaplan.empty:
        sheet_data["Anaplan"] = df_anaplan
        print(f"    Anaplan: {len(df_anaplan)} rows")
    else:
        print(f"    Anaplan: (empty)")

    output_path = os.path.join(output_dir, "PORB_MASTER_All_Programs.xlsx")
    write_consolidated_file(output_path, sheet_data, "MASTER")


def main():
    args = parse_args()
    porb_dir = os.path.abspath(args.input_dir)

    print("=" * 60)
    print("PORB Consolidation Script v3")
    print("  New Draft_PORB format with SP codes")
    print("=" * 60)
    print(f"\nInput directory: {porb_dir}")
    print(f"Output directory: {OUTPUT_DIR}")

    if not os.path.isdir(porb_dir):
        print(f"\nERROR: Input directory does not exist: {porb_dir}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    files = sorted(
        [
            f
            for f in os.listdir(porb_dir)
            if f.endswith(".xlsx") and "_PORB_" in f and re.search(r"_SP\d{2}_", f)
        ]
    )

    if not files:
        print("\nNo *_PORB_SP##_*.xlsx files found in the input directory.")
        print("Expected pattern: <Status>_PORB_<SPCode>_<Center>.xlsx")
        print("Examples: Draft_PORB_SP01_AfricaRice.xlsx, Pending_PORB_SP07_IFPRI.xlsx")
        sys.exit(1)

    print(f"\nFound {len(files)} PORB file(s)")

    programs_found = sorted(
        set(
            extract_program_center(f)[0]
            for f in files
            if extract_program_center(f)[0]
        )
    )
    print(f"Programs ({len(programs_found)}): {', '.join(programs_found)}")

    # Per-program summaries
    for program in programs_found:
        create_program_summary(program, files, porb_dir, OUTPUT_DIR)

    # Master file with all programs
    create_master_file(files, porb_dir, OUTPUT_DIR)

    print("\n" + "=" * 60)
    print("Consolidation complete!")
    print(f"Output files are in: {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
