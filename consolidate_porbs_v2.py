#!/usr/bin/env python3
"""
PORB Excel File Consolidation Script v2
- Fixed: Handles merged cells by forward-filling HLO and AOW columns
- Creates per-program summary files
- Creates master file with all programs
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PORB_DIR = os.path.join(SCRIPT_DIR, "PORBs")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "SummaryPORBs")

PROGRAM_NAMES = [
    "Breeding for Tomorrow",
    "Sustainable Farming",
    "Sustainable Animal and Aquatic Foods",
    "Multifunctional Landscapes",
    "Better Diets and Nutrition",
    "Climate Action",
    "Policy Innovations",
    "Food Frontiers and Security",
    "Scaling for Impact",
    "Gender Equality and Inclusion",
    "Capacity Sharing",
    "Digital Transformation",
    "Genebank",
]

# Sheets to consolidate
DATA_SHEETS = ['Cross-Cutting', 'HLO', 'Partner', 'W3-Bilateral projects', 'MELIA']

# Columns that might have merged cells and need forward-fill
MERGE_FILL_COLS = {
    'HLO': ['AOW', 'High Level Output'],
    'Partner': ['AOW'],
    'W3-Bilateral projects': ['AOW', 'Project title'],
    'Cross-Cutting': ['AOW'],
    'MELIA': ['AOW', 'MELIA study'],
}

# HLO sheet has a two-row header - define the proper column names
HLO_COLUMNS = [
    'AOW',
    'High Level Output',
    'KPI Description',
    'KPI Type',
    'KPI Geographic Location',
    'KPI Target',
    'KPI Budget (USD)',
    'Total Budget (USD)',
    'Budget Assumptions'
]

def extract_program_center(filename):
    """Extract program name and center name from filename"""
    basename = filename.replace('.xlsx', '')
    if not basename.startswith('Planning_'):
        return None, None
    rest = basename[len('Planning_'):]
    for program in sorted(PROGRAM_NAMES, key=len, reverse=True):
        if rest.startswith(program + '_'):
            center = rest[len(program) + 1:]
            return program, center
    return None, None

def clean_dataframe(df, sheet_name):
    """Clean dataframe - handle merged cells, remove headers and subtotals"""
    if df.empty:
        return df

    # Forward-fill merged cell columns
    if sheet_name in MERGE_FILL_COLS:
        for col in MERGE_FILL_COLS[sheet_name]:
            if col in df.columns:
                df[col] = df[col].ffill()

    # Remove rows that are completely empty
    df = df.dropna(how='all')

    # Remove subtotal rows
    for col in df.columns:
        if df[col].dtype == object:
            mask = df[col].astype(str).str.contains('subtotal|Subtotal|SUBTOTAL', na=False)
            df = df[~mask]

    return df.reset_index(drop=True)

def read_hlo_sheet(filepath):
    """Read HLO sheet with special handling for two-row headers"""
    # Read raw data without headers
    df = pd.read_excel(filepath, sheet_name='HLO', header=None)

    # Skip the first two header rows
    df = df.iloc[2:].reset_index(drop=True)

    # Assign proper column names (handle cases where file might have different column count)
    if len(df.columns) == len(HLO_COLUMNS):
        df.columns = HLO_COLUMNS
    elif len(df.columns) > len(HLO_COLUMNS):
        # More columns than expected - use HLO_COLUMNS for first N, keep rest as Unnamed
        df.columns = HLO_COLUMNS + [f'Extra_{i}' for i in range(len(df.columns) - len(HLO_COLUMNS))]
    else:
        # Fewer columns - use what we have
        df.columns = HLO_COLUMNS[:len(df.columns)]

    return df

def read_and_consolidate_sheet(files, sheet_name, porb_dir):
    """Read a specific sheet from all files and consolidate"""
    all_data = []

    for filename in files:
        program, center = extract_program_center(filename)
        if not program:
            continue

        filepath = os.path.join(porb_dir, filename)
        try:
            # Special handling for HLO sheet with two-row headers
            if sheet_name == 'HLO':
                df = read_hlo_sheet(filepath)
            else:
                df = pd.read_excel(filepath, sheet_name=sheet_name)

            df = clean_dataframe(df, sheet_name)

            if not df.empty:
                df.insert(0, 'Center', center)
                df.insert(0, 'Program/Accelerator', program)
                all_data.append(df)
        except Exception as e:
            print(f"  Warning: Error reading {sheet_name} from {filename}: {e}")

    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        return combined
    return pd.DataFrame()

def read_anaplan_sheet(files, porb_dir):
    """Read and consolidate Anaplan sheets"""
    all_data = []

    for filename in files:
        program, center = extract_program_center(filename)
        if not program:
            continue

        filepath = os.path.join(porb_dir, filename)
        try:
            df = pd.read_excel(filepath, sheet_name='Anaplan')
            df = df.dropna(how='all')

            if 'Main Accounts' in df.columns:
                mask = df['Main Accounts'].astype(str).str.contains('Subtotal|subtotal', na=False)
                df = df[~mask]

            if not df.empty:
                df.insert(0, 'Center', center)
                df.insert(0, 'Program/Accelerator', program)
                all_data.append(df)
        except Exception as e:
            print(f"  Warning: Error reading Anaplan from {filename}: {e}")

    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        return combined
    return pd.DataFrame()

def style_header(ws, num_cols):
    """Apply professional styling to header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_adjust_columns(ws, df):
    """Auto-adjust column widths"""
    for idx, col in enumerate(df.columns, 1):
        max_length = len(str(col))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max_length + 2

def write_consolidated_file(output_path, sheet_data_dict, title_prefix=""):
    """Write consolidated data to Excel file with multiple sheets"""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in sheet_data_dict.items():
        if df.empty:
            continue

        clean_name = sheet_name[:31].replace('/', '-')
        ws = wb.create_sheet(title=clean_name)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1:
                    cell.font = Font(name="Arial", size=10)

        style_header(ws, len(df.columns))
        auto_adjust_columns(ws, df)
        ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"  Saved: {output_path}")

def create_program_summary(program, files, porb_dir, output_dir):
    """Create summary file for a single program"""
    program_files = [f for f in files if extract_program_center(f)[0] == program]

    if not program_files:
        print(f"  No files found for {program}")
        return

    print(f"\nProcessing {program} ({len(program_files)} centers)...")

    sheet_data = {}

    for sheet_name in DATA_SHEETS:
        df = read_and_consolidate_sheet(program_files, sheet_name, porb_dir)
        if not df.empty:
            if 'Program/Accelerator' in df.columns:
                df = df.drop(columns=['Program/Accelerator'])
            sheet_data[sheet_name] = df
            print(f"    {sheet_name}: {len(df)} rows")

    df_anaplan = read_anaplan_sheet(program_files, porb_dir)
    if not df_anaplan.empty:
        if 'Program/Accelerator' in df_anaplan.columns:
            df_anaplan = df_anaplan.drop(columns=['Program/Accelerator'])
        sheet_data['Anaplan'] = df_anaplan
        print(f"    Anaplan: {len(df_anaplan)} rows")

    output_path = os.path.join(output_dir, f"PORB_Consolidated_{program}.xlsx")
    write_consolidated_file(output_path, sheet_data, program)


def create_master_file(files, porb_dir, output_dir):
    """Create master file with all programs combined"""
    print("\nCreating MASTER file (all programs)...")

    sheet_data = {}

    for sheet_name in DATA_SHEETS:
        df = read_and_consolidate_sheet(files, sheet_name, porb_dir)
        if not df.empty:
            sheet_data[sheet_name] = df
            print(f"    {sheet_name}: {len(df)} rows")

    df_anaplan = read_anaplan_sheet(files, porb_dir)
    if not df_anaplan.empty:
        sheet_data['Anaplan'] = df_anaplan
        print(f"    Anaplan: {len(df_anaplan)} rows")

    output_path = os.path.join(output_dir, "PORB_MASTER_All_Programs.xlsx")
    write_consolidated_file(output_path, sheet_data, "MASTER")

def main():
    print("=" * 60)
    print("PORB Consolidation Script v2 (with merged cell handling)")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    files = sorted([f for f in os.listdir(PORB_DIR) if f.endswith('.xlsx') and f.startswith('Planning_')])
    print(f"\nFound {len(files)} PORB files")

    programs = sorted(set(extract_program_center(f)[0] for f in files if extract_program_center(f)[0]))
    print(f"Programs: {', '.join(programs)}")

    for program in programs:
        create_program_summary(program, files, PORB_DIR, OUTPUT_DIR)

    create_master_file(files, PORB_DIR, OUTPUT_DIR)

    print("\n" + "=" * 60)
    print("Consolidation complete!")
    print("=" * 60)

if __name__ == "__main__":
    main()
